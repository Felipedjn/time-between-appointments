import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import openpyxl
import sys

def open_file(output_file):
    if output_file:
        resposta = messagebox.askyesno("Abrir Arquivo", "Deseja abrir o arquivo?")
        if resposta:
            # Aqui você pode adicionar a lógica para abrir o arquivo
            os.system(f'start excel "{output_file}"')
            print(f"Arquivo aberto: {output_file}")
        else:
            print("Arquivo não aberto.")
    else:
        print("Nenhum arquivo selecionado.")

def save(dados, output_file, sheet_names=['DADOS']):
    if output_file:
        sucess = False

        while not sucess:
            try:
                save_file(dados, output_file, sheet_names)
                sucess = True
                
            except Exception as e:
                messagebox.showerror("Erro", f"Ocorreu um erro: {e}\nNão foi possível salvar o arquivo, verifique se ele está aberto")
                output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")], initialfile=os.path.splitext(os.path.basename(selected_file))[0])

def save_file(dfs, path, sheet_names):
    writer = pd.ExcelWriter(path, engine='openpyxl')

    for df, s_name in zip(dfs, sheet_names):
        print(s_name)
        df.to_excel(writer, index=False, sheet_name=s_name)

        # Calcula a diferença de tempo com base na primeira ocorrência
        df['DIFERENCA_DE_TEMPO'] = df.groupby(['DADO', 'IP'])['HORARIO'].diff().fillna(pd.Timedelta(seconds=0))

        count_ = df['SEGMENTACAO'].value_counts().reset_index()
        count_.columns = ['SEGMENTACAO', 'QTD']

        total_querys = len(df)
        total = pd.Series({'SEGMENTACAO': 'Total de consultas', 'QTD': total_querys})

        count_ = count_.append(total, ignore_index=True)

        # Reordenando as linhas
        nova_ordem = ['Somente 1 consulta','Primeira consulta','Até 1 min', 'De 1 a 5 min', 'De 5 a 15 min',
                    'De 15 min a 1h', 'De 1h a 12h', 'De 12h a 24h', 'Acima de 24h', 'Total de consultas']

        count_['SEGMENTACAO'] = pd.Categorical(count_['SEGMENTACAO'], categories=nova_ordem, ordered=True)
        count_ = count_.sort_values('SEGMENTACAO')

        count_.to_excel(writer, index=False, sheet_name=f'CONTAGEM {s_name}')

    writer.close()

    writer = openpyxl.load_workbook(path)

    sheet_contagem = writer[f'CONTAGEM {s_name}']

    # Definir estilo para o cabeçalho
    header_style = openpyxl.styles.NamedStyle(
        name="header_style",
        font=openpyxl.styles.Font(color="00FFFFFF", bold=True),  # Cor branca e negrito
        fill=openpyxl.styles.PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),  # Cor vermelha
        border=openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin')),  # Borda inferior fina
        alignment=openpyxl.styles.Alignment(horizontal="center", vertical="center")  # Alinhamento centralizado
    )

    # Aplicar o estilo ao cabeçalho
    for cell in sheet_contagem['A1:B1']:  # Supondo que o cabeçalho esteja nas colunas A e B
        for col in cell:
            col.style = header_style

    for s_name in sheet_names:
        sheet_dados = writer[s_name]
        
        # Aplicar o estilo ao cabeçalho
        for cell in sheet_dados['A1:E1']:  # Supondo que o cabeçalho esteja nas colunas A e B
            for col in cell:
                col.style = header_style

    # Iterar sobre ambas as abas
    for sheet_name in sheet_names + [f'CONTAGEM {s_name}' for s_name in sheet_names]:
        sheet_contagem = writer[sheet_name]

        if sheet_name.startswith('CONTAGEM'):
            # Aplicar o estilo ao cabeçalho da planilha de contagem
            for cell in sheet_contagem['A1:B1']:
                for col in cell:
                    col.style = header_style
                    
        # Adicionar bordas a todas as células com dados
        for row in sheet_contagem.iter_rows(min_row=2, max_row=sheet_contagem.max_row, min_col=1, max_col=sheet_contagem.max_column):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

        # Centralizar todas as células
        for row in sheet_contagem.iter_rows(min_row=1, max_row=sheet_contagem.max_row, min_col=1, max_col=sheet_contagem.max_column):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        # Ajustar o espaçamento na coluna
        for column_cells in sheet_contagem.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet_contagem.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # Salvar as alterações
    writer.save(path)

def select_file():
    global selected_file
    selected_file = filedialog.askopenfilename(filetypes=[("Todos os arquivos", "*.*")])
    file_name = selected_file.split('/')
    file_name = file_name[-1]
    tk_label_file.config(text=f'Arquivo: {file_name}')

def first_query_ip_on(dados):
    # Classifica as colunas na ordem respectiva
    dados = dados.sort_values(by=['IP', 'DADO', 'HORARIO'])
    dados['PRIMEIRA_CONSULTA'] = ~dados.duplicated(subset=['DADO'], keep='first')
    # Encontra o primeiro horário de cada dado
    dados['PRIMEIRO_HORARIO'] = dados.groupby(['DADO', 'IP'])['HORARIO'].transform('first')
    # Calcula a diferença de tempo entre a consulta atual e a primeira consulta
    dados['DIFERENCA_DE_TEMPO'] = dados['HORARIO'] - dados['PRIMEIRO_HORARIO']
    # Separa os dados que não têm mais de 1 ocorrência
    dados['UNICA_CONSULTA'] = ~dados[['DADO', 'IP']].duplicated(keep=False)

    return dados

def last_query_ip_on(dados):
    # Classifica as colunas na ordem respectiva
    dados = dados.sort_values(by=['IP', 'DADO', 'HORARIO'])
    # Separa a primeira ocorrencia daqueles dados duplicados
    dados['PRIMEIRA_CONSULTA'] = ~dados.duplicated(subset=['DADO'], keep='first')
    # Acha a diferença de tempo do agrupamento respectivo
    dados['DIFERENCA_DE_TEMPO'] = dados.groupby(['DADO', 'IP'])['HORARIO'].diff()
    # Separa os dados que não tem mais de 1 ocorrencia
    dados['UNICA_CONSULTA'] = ~dados[['DADO','IP']].duplicated(keep=False)

    return dados

def first_query_ip_off(dados):
    # Classifica as colunas na ordem respectiva
    dados = dados.sort_values(by=['DADO', 'HORARIO'])
    dados['PRIMEIRA_CONSULTA'] = ~dados.duplicated(subset=['DADO'], keep='first')
    # Encontra o primeiro horário de cada dado
    dados['PRIMEIRO_HORARIO'] = dados.groupby(['DADO'])['HORARIO'].transform('first')
    # Calcula a diferença de tempo entre a consulta atual e a primeira consulta
    dados['DIFERENCA_DE_TEMPO'] = dados['HORARIO'] - dados['PRIMEIRO_HORARIO']
    # Separa os dados que não têm mais de 1 ocorrência
    dados['UNICA_CONSULTA'] = ~dados['DADO'].duplicated(keep=False)

    return dados

def last_query_ip_off(dados):
    # Classifica as colunas na ordem respectiva
    dados = dados.sort_values(by=['DADO', 'HORARIO'])
    # Separa a primeira ocorrencia daqueles dados duplicados
    dados['PRIMEIRA_CONSULTA'] = ~dados.duplicated(subset=['DADO'], keep='first')
    # Acha a diferença de tempo do agrupamento respectivo
    dados['DIFERENCA_DE_TEMPO'] = dados.groupby(['DADO'])['HORARIO'].diff()
    # Separa os dados que não tem mais de 1 ocorrencia
    dados['UNICA_CONSULTA'] = ~dados['DADO'].duplicated(keep=False)

    return dados

def set_conditions(dados, qtd_lines):
        # Condições de tempo
        conditions = [
            (dados['UNICA_CONSULTA']), # Somente 1 consulta
            (dados['PRIMEIRA_CONSULTA']), # Primeira consulta
            (dados['DIFERENCA_DE_TEMPO'] <= '0 days 00:01:00'),  # Até 1 min
            ((dados['DIFERENCA_DE_TEMPO'] > '0 days 00:01:00') & (dados['DIFERENCA_DE_TEMPO'] <= '0 days 00:05:00')),  # De 1 min a 5 min
            ((dados['DIFERENCA_DE_TEMPO'] > '0 days 00:05:00') & (dados['DIFERENCA_DE_TEMPO'] <= '0 days 00:15:00')),  # De 5 min a 15min
            ((dados['DIFERENCA_DE_TEMPO'] > '0 days 00:15:00') & (dados['DIFERENCA_DE_TEMPO'] <= '0 days 01:00:00')),  # De 15 min a 1h
            ((dados['DIFERENCA_DE_TEMPO'] > '0 days 01:00:00') & (dados['DIFERENCA_DE_TEMPO'] <= '0 days 12:00:00')),  # De 1h a 12h
            ((dados['DIFERENCA_DE_TEMPO'] > '0 days 12:00:00') & (dados['DIFERENCA_DE_TEMPO'] <= '1 days 00:00:00')),  # De 12h a 24h
            (dados['DIFERENCA_DE_TEMPO'] > '1 days 00:00:00')  # Acima de 24h
        ]

        # Legenda das condições de tempo, respectivamente
        choices = ['Somente 1 consulta', 'Primeira consulta', 'Até 1 min', 'De 1 a 5 min', 'De 5 a 15 min', 'De 15 min a 1h', 'De 1h a 12h', 'De 12h a 24h', 'Acima de 24h']

        # Cria a coluna SEGMENTAÇÃO e adiciona as legendas
        dados['SEGMENTACAO'] = np.select(conditions, choices)

        # Dropa as colunas desnecessárias
        try:
            dados = dados.drop(columns=['PRIMEIRA_CONSULTA', 'DIFERENCA_DE_TEMPO', 'UNICA_CONSULTA', 'PRIMEIRO_HORARIO'])
        except:
            dados = dados.drop(columns=['PRIMEIRA_CONSULTA', 'DIFERENCA_DE_TEMPO', 'UNICA_CONSULTA'])

        # Substitui os valores que não conseguiu aplicar a condição (são sempre os de Primeira consulta)
        dados['SEGMENTACAO'] = dados['SEGMENTACAO'].replace(['0'], 'Primeira consulta')

        print(dados)

        return dados

def process_file():
    if selected_file:
        try:
            dados = pd.read_csv(selected_file, sep=';', encoding='utf-8')
        except:
            dados = pd.read_excel(selected_file)

        qtd_lines = len(dados)
        progress["maximum"] = 100

        dados['HORARIO'] = pd.to_datetime(dados['HORARIO'])
        dados['HORARIO'] = dados['HORARIO'].dt.strftime('%Y/%m/%d %H:%M:%S')

        # Converte a coluna Horario
        dados['HORARIO'] = pd.to_datetime(dados['HORARIO'], format='%Y/%m/%d %H:%M:%S')
        
        if var_first_query.get() and var_ipon.get() and var_last_query.get() and var_ipoff.get():
            dados_1 = first_query_ip_on(dados)
            dados_2 = last_query_ip_on(dados)
            dados_3 = first_query_ip_off(dados)
            dados_4 = last_query_ip_off(dados)

            progress["value"] = 50

            sheet_names = ['DADOS ON PC', 'DADOS ON UC', 'DADOS OFF PC', 'DADOS OFF UC']
            dados = [dados_1, dados_2, dados_3, dados_4]

            
            output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")], initialfile=os.path.splitext(os.path.basename(selected_file))[0])

            lista_dados = []
            for df in dados:
                dados_ = set_conditions(df, qtd_lines)
                lista_dados.append(dados_)

            save(lista_dados, output_file, sheet_names)

            progress["value"] = 100

            messagebox.showinfo("Concluído", "Arquivo salvo com sucesso!")
            open_file(output_file)

            progress["value"] = 0

        elif var_first_query.get() and var_ipon.get() and var_ipoff.get():
            dados_1 = first_query_ip_on(dados)
            dados_3 = first_query_ip_off(dados)

            progress["value"] = 50

            sheet_names = ['ON PC', 'OFF PC']
            dados = [dados_1, dados_3]

            
            output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")], initialfile=os.path.splitext(os.path.basename(selected_file))[0])

            lista_dados = []
            for df in dados:
                dados_ = set_conditions(df, qtd_lines)
                lista_dados.append(dados_)

            save(lista_dados, output_file, sheet_names)

            progress["value"] = 100

            messagebox.showinfo("Concluído", "Arquivo salvo com sucesso!")
            open_file(output_file)

            progress["value"] = 0

        elif var_last_query.get() and var_ipon.get() and var_ipoff.get():
            dados_2 = last_query_ip_on(dados)
            dados_4 = last_query_ip_off(dados)

            progress["value"] = 50

            sheet_names = ['ON UC', 'OFF UC']
            dados = [dados_2, dados_4]

            
            output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")], initialfile=os.path.splitext(os.path.basename(selected_file))[0])

            lista_dados = []
            for df in dados:
                dados_ = set_conditions(df, qtd_lines)
                lista_dados.append(dados_)

            save(lista_dados, output_file, sheet_names)

            progress["value"] = 100

            messagebox.showinfo("Concluído", "Arquivo salvo com sucesso!")
            open_file(output_file)

            progress["value"] = 0

        elif var_first_query.get() and var_last_query.get() and var_ipoff.get():
            dados_3 = first_query_ip_off(dados)
            dados_4 = last_query_ip_off(dados)

            progress["value"] = 50

            sheet_names = ['DADOS OFF PC', 'DADOS OFF UC']
            dados = [dados_3, dados_4]

            
            output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")], initialfile=os.path.splitext(os.path.basename(selected_file))[0])
            
            lista_dados = []
            for df in dados:
                dados_ = set_conditions(df, qtd_lines)
                lista_dados.append(dados_)

            save(lista_dados, output_file, sheet_names)

            progress["value"] = 100

            messagebox.showinfo("Concluído", "Arquivo salvo com sucesso!")
            open_file(output_file)

            progress["value"] = 0

        elif var_first_query.get() and var_ipon.get() and var_last_query.get():
            dados_1 = first_query_ip_on(dados)
            dados_2 = last_query_ip_on(dados)

            progress["value"] = 50

            sheet_names = ['DADOS ON PC', 'DADOS ON UC']
            dados = [dados_1, dados_2]

            
            output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")], initialfile=os.path.splitext(os.path.basename(selected_file))[0])

            lista_dados = []
            for df in dados:
                dados_ = set_conditions(df, qtd_lines)
                lista_dados.append(dados_)

            save(lista_dados, output_file, sheet_names)

            progress["value"] = 100

            messagebox.showinfo("Concluído", "Arquivo salvo com sucesso!")

            progress["value"] = 0

        else:
            if var_first_query.get() and var_ipon.get():
                print('PRIMEIRA CONSULTA - IP ON')
                dados = first_query_ip_on(dados)
                progress["value"] = 50
            elif var_last_query.get() and var_ipon.get():
                print('ULTIMA CONSULTA - IP ON')
                dados = last_query_ip_on(dados)
                progress["value"] = 50
            elif var_first_query.get() and var_ipoff.get():
                print('PRIMEIRA CONSULTA - IP OFF')
                dados = first_query_ip_off(dados)
                progress["value"] = 50
            elif var_last_query.get() and var_ipoff.get():
                print('ULTIMA CONSULTA - IP OFF')
                dados = last_query_ip_off(dados)
                progress["value"] = 50

            print(dados)

            dados = set_conditions(dados, qtd_lines)
            
            output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")], initialfile=os.path.splitext(os.path.basename(selected_file))[0])
            save([dados], output_file)

            progress["value"] = 100
            messagebox.showinfo("Concluído", "Arquivo salvo com sucesso!")
            open_file(output_file)
            
            progress["value"] = 0
            
    else:
        messagebox.showerror("Erro", "Nenhum arquivo selecionado")

        progress["value"] = 0


def resource_path(relative_path):    
    try:       
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

datafile = "./icon/icon.ico"

# Criando a janela principal
window = tk.Tk()
window.configure(bg='#0D0D0D')
window.title("Tempo Entre Consultas")
window.iconbitmap(default=resource_path(datafile))
window.geometry("350x190")

# Botões e rótulos
tk_button_select = tk.Button(window, text="Selecionar Arquivo", width=15, fg='white', command=select_file)
tk_button_process = tk.Button(window, text="Processar Arquivo", width=15, fg='white', command=process_file)
tk_label_file = tk.Label(window, text="Arquivo: Nenhum", fg='white')
# Variáveis para armazenar o estado dos botões
var_ipon = tk.BooleanVar(value=True)
var_ipoff = tk.BooleanVar()
var_last_query = tk.BooleanVar(value=True)
var_first_query = tk.BooleanVar()

# Cria os botões
botton_ipon = tk.Checkbutton(window, text="IP ON", variable=var_ipon)
botton_ipoff = tk.Checkbutton(window, text="IP OFF", variable=var_ipoff)
botton_last_query = tk.Checkbutton(window, text="ULTIMA CONSULTA", variable=var_last_query)
botton_first_query = tk.Checkbutton(window, text="PRIMEIRA CONSULTA", variable=var_first_query)

botton_ipon.pack()
botton_ipoff.pack()
botton_last_query.pack()
botton_first_query.pack()

botton_last_query.place(x=180, y=60)
botton_last_query.configure(activebackground='#0D0D0D', bg='#0D0D0D', fg='white', selectcolor='#8b0304')
botton_first_query.place(x=180, y=80)
botton_first_query.configure(activebackground='#0D0D0D', bg='#0D0D0D', fg='white', selectcolor='#8b0304')
botton_ipon.place(x=180, y=10)
botton_ipon.configure(activebackground='#0D0D0D', bg='#0D0D0D', fg='white', selectcolor='#8b0304')
botton_ipoff.place(x=180, y=30)
botton_ipoff.configure(activebackground='#0D0D0D', bg='#0D0D0D', fg='white', selectcolor='#8b0304')

tk_button_select.place(x=20, y=16)
tk_button_select.configure(bg='#8b0304')
tk_button_process.place(x=20, y=51)
tk_button_process.configure(bg='#8b0304')
tk_label_file.place(x=19, y=155)
tk_label_file.configure(bg='#0D0D0D')

# Cria uma barra de progress na interface gráfica.
progress = ttk.Progressbar(window, orient="horizontal", length=300, mode="determinate")
progress.place(x=20, y=125)

# Ajustar janela
window.update_idletasks()
width = window.winfo_width()
height = window.winfo_height()
x = (window.winfo_screenwidth() // 2) - (width // 2)
y = (window.winfo_screenheight() // 2) - (height // 2)
window.geometry(f"{width}x{height}+{x}+{y}")

# Inicia o loop principal do tkinter
window.mainloop()
